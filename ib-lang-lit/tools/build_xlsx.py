#!/usr/bin/env python3
"""Assemble the final IB Lang&Lit v3 database: enriched originals + new activities."""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import content_texts, content_craft, content_support, content_extra, content_final

COLUMNS = ['Activity ID', 'Assessment', 'Category', 'Skill Focus', 'Activity Type',
           'Activity Title', 'Activity Description', 'Time (min)', 'IB Criterion',
           'Style Preference', 'Text Type(s)', 'Text / Work Focus', 'Difficulty',
           'Location', 'Format', 'Session Use', 'Exam Proximity', 'Materials',
           'Energy Level', 'Support Level', 'Helps With', 'Why It Helps',
           'Success Criteria', 'If You Are Stuck']

def load_all():
    with open('enriched.json') as f:
        old = json.load(f)
    new = []
    for m in (content_texts, content_craft, content_support, content_extra, content_final):
        new += m.all_rows()
    # de-collide titles between old and new sets
    old_titles = {r['Activity Title'] for r in old}
    for r in new:
        if r['Activity Title'] in old_titles:
            r['Activity Title'] = r['Activity Title'] + ' (Deep Dive)'
    # difficulty upgrades for heavy sessions
    for r in new:
        if r['Time (min)'] >= 60 and r['Session Use'] in ('Exam Practice', 'Core Activity'):
            r['Difficulty'] = 'Advanced'
    # normalize session labels
    for r in old + new:
        if r['Session Use'] == 'IO Prep':
            r['Session Use'] = 'IO Preparation'
    return old, new

def assign_ids(rows):
    counters = {}
    for r in rows:
        parts = [a.strip() for a in str(r['Assessment']).split(',') if a.strip()]
        if len(parts) > 1:
            prefix = 'X'          # cross-assessment
        elif parts and parts[0] == 'Paper 1':
            prefix = 'P1'
        elif parts and parts[0] == 'Paper 2':
            prefix = 'P2'
        elif parts and parts[0] == 'Individual Oral':
            prefix = 'IO'
        else:
            prefix = 'GEN'
        counters[prefix] = counters.get(prefix, 0) + 1
        r['Activity ID'] = f'{prefix}-{counters[prefix]:04d}'
    return rows

def write_workbook(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Activities'

    navy = '1E3A5F'
    header_fill = PatternFill('solid', fgColor=navy)
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    body_font = Font(name='Arial', size=10)
    thin = Side(style='thin', color='DDD8CE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(COLUMNS)
    for c, _ in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = border

    for r in rows:
        ws.append([r.get(col, '') for col in COLUMNS])

    wrap_cols = {'Activity Description', 'Why It Helps', 'Success Criteria', 'If You Are Stuck'}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            colname = COLUMNS[cell.column - 1]
            cell.alignment = Alignment(vertical='top',
                                       wrap_text=(colname in wrap_cols))

    widths = {'Activity ID': 10, 'Assessment': 22, 'Category': 24, 'Skill Focus': 24,
              'Activity Type': 16, 'Activity Title': 34, 'Activity Description': 70,
              'Time (min)': 9, 'IB Criterion': 11, 'Style Preference': 26,
              'Text Type(s)': 24, 'Text / Work Focus': 26, 'Difficulty': 13,
              'Location': 16, 'Format': 11, 'Session Use': 14, 'Exam Proximity': 20,
              'Materials': 24, 'Energy Level': 22, 'Support Level': 30,
              'Helps With': 32, 'Why It Helps': 55, 'Success Criteria': 48,
              'If You Are Stuck': 48}
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{ws.max_row}'

    # ── Read Me tab ────────────────────────────────────────────────────
    rm = wb.create_sheet('Read Me')
    rm.sheet_view.showGridLines = False
    title_font = Font(name='Arial', size=16, bold=True, color=navy)
    h_font = Font(name='Arial', size=11, bold=True, color=navy)
    b_font = Font(name='Arial', size=10)
    rm.column_dimensions['A'].width = 30
    rm.column_dimensions['B'].width = 100

    def put(row, a, b='', af=None, bf=None):
        rm.cell(row=row, column=1, value=a).font = af or b_font
        if b:
            cell = rm.cell(row=row, column=2, value=b)
            cell.font = bf or b_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        return row + 1

    n = {}
    for r in rows:
        for a in str(r['Assessment']).split(','):
            a = a.strip()
            if a:
                n[a] = n.get(a, 0) + 1
    multi = sum(1 for r in rows if ',' in str(r['Assessment']))

    row = put(1, 'IB English A: Language & Literature — Master Activity Database v3', af=title_font)
    row = put(row, '')
    row = put(row, 'What this is',
              f'{len(rows):,} practice activities for Paper 1, Paper 2, the Individual Oral, and general study skills. '
              f'This sheet powers the Activity Explorer web app: the app reads the "Activities" tab directly (no Apps Script needed).', af=h_font)
    row = put(row, 'Coverage',
              f"Paper 1: {n.get('Paper 1',0):,} · Paper 2: {n.get('Paper 2',0):,} · Individual Oral: {n.get('Individual Oral',0):,} · "
              f"General: {n.get('General',0):,} · Cross-assessment activities (count in 2+ groups): {multi:,}", af=h_font)
    row = put(row, '')
    row = put(row, 'How to update the app data', af=h_font)
    row = put(row, '1.', 'In Google Sheets: File → Import → Upload → this file → "Replace spreadsheet" (or replace just the Activities tab).')
    row = put(row, '2.', 'Keep the tab named exactly "Activities" and keep the header row as row 1.')
    row = put(row, '3.', 'Keep sharing set to "Anyone with the link — Viewer". The app fetches fresh data on every page load.')
    row = put(row, '4.', 'The app matches columns BY HEADER NAME, so you may reorder columns or add your own — just don\'t rename the headers below. '
                          'You can edit any cell, add rows, or delete rows freely.')
    row = put(row, '')
    row = put(row, 'Column guide (allowed values)', af=h_font)
    guide = [
      ('Activity ID', 'Unique reference (P1-/P2-/IO-/GEN-/X- prefix; X = cross-assessment). Shown on cards.'),
      ('Assessment', 'One or more of: Paper 1, Individual Oral, Paper 2, General (comma-separated). Activities listing 2+ appear under each and in the Cross-Assessment filter.'),
      ('Category', 'Broad grouping shown in the Category filter (e.g. "Paper 2: The World\'s Wife", "Cross-Assessment Bridge", "Scaffolded Support").'),
      ('Skill Focus', 'The specific skill being trained. Populates the Skill Focus filter.'),
      ('Activity Type', 'Task style (Skill Drill, Analytical Activity, Oral Dialogue, Exam Practice, IO Prep, …).'),
      ('Activity Title', 'Short name shown as the card title.'),
      ('Activity Description', 'The full student-facing task. 2–5 sentences, self-contained.'),
      ('Time (min)', 'Whole minutes. The app rolls these into time bands (Quick ≤5, Short ≤15, Medium ≤30, Long ≤60, Extended 60+).'),
      ('IB Criterion', 'A, B, C, D — comma-separated for multiple.'),
      ('Style Preference', 'Any of: Visual, Spatial, Auditory, Reading, Writing, Kinesthetic, Verbal, Linguistic, Logical, Analytical, Social/Interpersonal, Solitary, Intrapersonal, Naturalist, Multimodal (comma-separated). Legacy abbreviations (V, R/W, Sol…) are also understood by the app.'),
      ('Text Type(s)', 'Text types involved (Advertisement, Poems, Graphic Novel, …) or "Any".'),
      ('Text / Work Focus', 'Studied works involved: William Blake, Thi Bui, Banksy, Barbara Kruger, Ugur Gallenkuş, The World\'s Wife, A Doll\'s House (comma-separated). Blank = not work-specific.'),
      ('Difficulty', 'Foundational, Developing, or Advanced.'),
      ('Location', 'At Home, In Class, On the Go, Out & About, During Screen Time, Anywhere, At Home or In Class.'),
      ('Format', 'Individual, Pair, Group (or "Flexible").'),
      ('Session Use', 'Warm-Up, Core Activity, Exam Practice, Writing Task, Reflection, Cool-Down, IO Preparation.'),
      ('Exam Proximity', 'Any Time, Months Away, Weeks Away, Final Week, Night Before (comma-separated when several apply).'),
      ('Materials', 'What is needed: Pen & Paper, Device, Timer, Flashcards, Recording Device, Printed Text or Screen, or "None — just your eyes and brain".'),
      ('Energy Level', 'Low — fine when tired / Medium — normal study energy / High — needs full focus.'),
      ('Support Level', 'Highly Scaffolded — steps are laid out for you / Some Scaffolding — guided but open / Independent — you drive the thinking.'),
      ('Helps With', 'Struggle tags for the "I want help with…" filter: Getting Started, Analysis Depth, Structure & Planning, Time Management, Quotations & Evidence, Comparative Linking, Speaking Confidence, Vocabulary & Terminology, Memory & Recall, Focus & Motivation, Fresh Ideas & Interpretation, Exam Confidence & Calm.'),
      ('Why It Helps', 'Shown on the card: the payoff, in student language.'),
      ('Success Criteria', 'Shown on the card: "you\'ll know it worked when…".'),
      ('If You Are Stuck', 'Shown on the card: a concrete first move for a struggling student.'),
    ]
    for name, desc in guide:
        row = put(row, name, desc, af=Font(name='Arial', size=10, bold=True))
    row = put(row, '')
    row = put(row, 'Provenance', 'v3 (2026): 1,926 original activities enriched with the new columns (345 clipped descriptions repaired), '
                    'plus 1,399 new activities including text-specific drills for all studied works and cross-assessment bridge activities. '
                    'Original Skill labels that named learner styles, everyday contexts, or artists were moved into the proper dedicated columns.')
    wb.save(path)

def main():
    old, new = load_all()
    rows = assign_ids(old + new)
    df = pd.DataFrame(rows)[COLUMNS]
    # integrity checks
    assert df['Activity ID'].is_unique
    assert df['Activity Title'].notna().all() and df['Activity Description'].notna().all()
    assert (df['Time (min)'] > 0).all()
    write_workbook(rows, 'IB_LangLit_Master_Activity_Database_v3.xlsx')
    df.to_csv('IB_LangLit_Master_Activity_Database_v3.csv', index=False)
    df.to_json('combined.json', orient='records')
    print(f'Wrote {len(rows):,} activities')
    print('\nAssessment tags:')
    tags = {}
    for a in df['Assessment']:
        for t in str(a).split(','):
            t = t.strip()
            tags[t] = tags.get(t, 0) + 1
    for k, v in sorted(tags.items(), key=lambda x: -x[1]): print(f'  {v:5,}  {k}')
    print('\nID prefixes:', df['Activity ID'].str.split('-').str[0].value_counts().to_dict())
    print('Cross-assessment rows:', int((df['Assessment'].str.contains(',')).sum()))
    print('Session Use:', df['Session Use'].value_counts().to_dict())
    print('Exam Proximity buckets:', df['Exam Proximity'].value_counts().head(8).to_dict())

if __name__ == '__main__':
    main()
